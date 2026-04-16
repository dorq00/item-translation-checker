"""
translation_checker — run.py
=============================
Drop your file in    invoices/incoming/
Double-click         START_CHECKER.bat
Done — output lands in output/YYYY-MM/

FOLDER STRUCTURE:
  translation_checker/
  ├── run.py                   ← this file
  ├── START_CHECKER.bat        ← double-click launcher
  ├── db/                      ← translation history (all .xlsx files are loaded)
  │   ├── Historical_Translation.xlsx   (seed — curated base)
  │   ├── MEINV00084440.xlsx            (past invoices moved here after review)
  │   └── ...
  ├── invoices/
  │   └── incoming/            ← drop your file here before running
  └── output/
      └── 2026-04/             ← auto-created, organized by month

TWO MODES — detected automatically:

  PO MODE  (file has "HS Code" sheet — LG Korea PO Master)
    The PO is a master catalog of all upcoming items, no translations included.
    Output: 3 sheets
      Sheet 1 — Coverage        (all items — overview)
      Sheet 2 — Need Translation (NEW items — paste to Claude)
      Sheet 3 — Ready           (items already in DB — translation auto-filled)

  MEINV MODE  (file has "DATA DETAILS" sheet — raw MEINV from LG portal)
    The invoice carries LG Korea's French designation.  Compare vs DB.
    Output: 3 sheets
      Sheet 1 — Status Check    (all rows — full audit)
      Sheet 2 — Changed         (translation conflicts — need decision)
      Sheet 3 — New Items       (not in DB — need translation)

HOW THE DB WORKS:
  All .xlsx files in db/ are loaded and merged at runtime.
  Files are sorted by name — later file wins on duplicate Item.
  To add translations: save a file with columns class|Item|Designation|translation
  and drop it in db/.  No other step needed.
"""

import os
import sys
from datetime import datetime
from pathlib import Path

# ── Dependencies ──────────────────────────────────────────────────────────────

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas not installed.  Run:  pip install pandas openpyxl")
    try:
        input("\nPress Enter to close...")
    except (EOFError, KeyboardInterrupt):
        pass
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    try:
        input("\nPress Enter to close...")
    except (EOFError, KeyboardInterrupt):
        pass
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────

ROOT       = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
DB_DIR     = ROOT / "db"
INCOMING   = ROOT / "invoices" / "incoming"
OUTPUT_DIR = ROOT / "output"

# ── Status labels ─────────────────────────────────────────────────────────────

NO_CHANGE = "✅ NO CHANGE"
CHANGED   = "🛑 CHANGED"
NEW_ITEM  = "🆕 NEW ITEM"
IN_DB     = "📋 IN DB"

# ── Colors ────────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

GREEN_FILL  = _fill("C6EFCE")
RED_FILL    = _fill("FFC7CE")
ORANGE_FILL = _fill("FFEB9C")
GREY_FILL   = _fill("F2F2F2")

GREEN_FONT  = Font(color="006100", bold=True)
RED_FONT    = Font(color="9C0006", bold=True)
ORANGE_FONT = Font(color="9C5700", bold=True)
WHITE_FONT  = Font(color="FFFFFF", bold=True)

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

STATUS_STYLE = {
    NO_CHANGE: (GREEN_FILL,  GREEN_FONT),
    IN_DB:     (GREEN_FILL,  GREEN_FONT),
    CHANGED:   (RED_FILL,    RED_FONT),
    NEW_ITEM:  (ORANGE_FILL, ORANGE_FONT),
}

# ── Loaders ───────────────────────────────────────────────────────────────────

def _clean_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none") else s


# LG Korea division codes → Algeria customs class codes
_DIV_MAP = {
    "W/M": "WM", "RAC": "AC", "CAC": "AC",
    "REF": "RF", "LTV": "TV", "MNT": "TV",
}


def _detect_mode(path: Path) -> str:
    """Return 'PO', 'MEINV', or 'flat' based on sheet names."""
    xl = pd.ExcelFile(path)
    if "DATA DETAILS" in xl.sheet_names:
        return "MEINV"
    if "SPM_Recommend" in xl.sheet_names:
        return "PO"
    if "HS Code" in xl.sheet_names:
        return "PO"
    return "flat"


def _parse_invoice_file(path: Path) -> pd.DataFrame:
    """
    Parse a single file into canonical columns:
      class, Item, translation, Designation, hs_code
    translation is "" when the file has no translation column (PO format).
    hs_code is "" for MEINV/flat formats (not present in those files).

    Handles three formats:
      - Raw MEINV from LG Korea portal (has DATA DETAILS sheet):
          Product → class | Item → Item | Designation → translation | Item Desc → Designation
      - PO Master (has HS Code sheet):
          Div Name → class | Part Number → Item | Product Description → Designation
          Customs Code → hs_code | (no translation — filled as "")
      - Flat/translated format (first sheet):
          class | Item | translation | Designation
    """
    xl = pd.ExcelFile(path)

    if "DATA DETAILS" in xl.sheet_names:
        df = xl.parse("DATA DETAILS", dtype=str).fillna("")
        df.columns = [str(c).strip() for c in df.columns]
        df = df.rename(columns={
            "Product":     "class",
            "Item":        "Item",
            "Designation": "translation",
            "Item Desc":   "Designation",
        })
        df["hs_code"] = ""

    elif "SPM_Recommend" in xl.sheet_names:
        # Header is on row 9 (0-indexed), real data starts after
        df = xl.parse("SPM_Recommend", header=9, dtype=str).fillna("")
        df.columns = [str(c).strip() for c in df.columns]
        df = df.rename(columns={
            "Host Part Id": "Item",
            "Part Name":    "Designation",
            "Designation":  "translation",
            "Div Name":     "class",
        })
        df["class"] = df["class"].map(lambda v: _DIV_MAP.get(v, v))
        df["hs_code"] = ""
        # Only rows with a French Designation are actual PO items.
        # The rest are planning/forecast rows — exclude them entirely.
        df = df[df["translation"].str.strip() != ""]

    elif "HS Code" in xl.sheet_names:
        df = xl.parse("HS Code", dtype=str).fillna("")
        df.columns = [str(c).strip() for c in df.columns]
        df = df.rename(columns={
            "Div Name":            "class",
            "Part Number":         "Item",
            "Product Description": "Designation",
            "Customs Code":        "hs_code",
        })
        df["class"] = df["class"].map(lambda v: _DIV_MAP.get(v, v))
        df["translation"] = ""
        if "hs_code" not in df.columns:
            df["hs_code"] = ""

    else:
        df = xl.parse(xl.sheet_names[0], dtype=str).fillna("")
        df.columns = [str(c).strip() for c in df.columns]
        col_map = {}
        for c in df.columns:
            lc = c.lower()
            if lc == "class":         col_map[c] = "class"
            elif lc == "item":        col_map[c] = "Item"
            elif lc == "translation": col_map[c] = "translation"
            elif lc == "designation": col_map[c] = "Designation"
            elif lc in ("hs_code", "hs code", "customs code"): col_map[c] = "hs_code"
        df = df.rename(columns=col_map)
        if "translation" not in df.columns:
            df["translation"] = ""
        if "hs_code" not in df.columns:
            df["hs_code"] = ""

    required = {"class", "Item", "Designation"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"{path.name}: missing columns {missing}\n"
            f"  Found: {list(df.columns)}"
        )

    df = df[["class", "Item", "translation", "Designation", "hs_code"]].copy()
    for col in df.columns:
        df[col] = df[col].apply(_clean_str)
    return df[df["Item"].str.strip() != ""].reset_index(drop=True)


def load_invoice(path: Path):
    """Load and parse the incoming file.  Returns (df, mode)."""
    mode = _detect_mode(path)
    df = _parse_invoice_file(path)
    print(f"  File     : {len(df)} rows  ({path.name})")
    print(f"  Mode     : {mode}")
    return df, mode


def load_db() -> pd.DataFrame:
    """
    Load translation history from all .xlsx files in db/.
    Files are sorted by name (MEINV numbers are chronological),
    so drop_duplicates keep='last' keeps the most recent translation per Item.
    """
    db_files = sorted(
        [f for f in DB_DIR.glob("*.xlsx") if not f.name.startswith("~$")],
        key=lambda f: f.name,
    )
    if not db_files:
        return pd.DataFrame(columns=["class", "Item", "translation", "Designation"])

    frames = []
    for f in db_files:
        try:
            frames.append(_parse_invoice_file(f))
        except Exception as e:
            print(f"  WARNING: skipping {f.name} — {e}")

    if not frames:
        return pd.DataFrame(columns=["class", "Item", "translation", "Designation"])

    df = pd.concat(frames, ignore_index=True)
    df = df[df["Item"].str.strip() != ""]
    # Only keep rows that actually have a translation (DB rows without one are useless as lookup)
    df = df[df["translation"].str.strip() != ""]
    df = df.drop_duplicates(subset="Item", keep="last").reset_index(drop=True)
    print(f"  DB       : {len(df)} unique items  ({len(db_files)} files)")
    return df

# ── Check logic ───────────────────────────────────────────────────────────────

def classify(df_invoice: pd.DataFrame, df_db: pd.DataFrame) -> pd.DataFrame:
    """
    LEFT JOIN invoice vs DB on Item.
    Adds db_translation column and Status.
    """
    db_lookup = df_db.set_index("Item")[["translation"]].rename(
        columns={"translation": "db_translation"}
    )
    df = df_invoice.merge(db_lookup, on="Item", how="left")
    df["db_translation"] = df["db_translation"].fillna("")
    if "hs_code" not in df.columns:
        df["hs_code"] = ""

    def _status(row):
        if row["db_translation"] == "":
            return NEW_ITEM
        if row["translation"] == "":          # PO format — no translation submitted
            return IN_DB
        if row["translation"].strip().upper() == row["db_translation"].strip().upper():
            return NO_CHANGE
        return CHANGED

    df["Status"] = df.apply(_status, axis=1)
    return df

# ── Excel output helpers ──────────────────────────────────────────────────────

def _write_header(ws, columns, fill):
    for ci, col in enumerate(columns, 1):
        c = ws.cell(1, ci, col)
        c.fill = fill
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
    ws.row_dimensions[1].height = 20


def _write_rows(ws, df):
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, "" if pd.isna(val) else val)
            c.alignment = Alignment(vertical="center")
            c.border = THIN
            if ri % 2 == 0:
                c.fill = GREY_FILL


def _color_status_col(ws, df, col_name):
    if col_name not in df.columns:
        return
    ci = list(df.columns).index(col_name) + 1
    for ri, val in enumerate(df[col_name], 2):
        style = STATUS_STYLE.get(str(val))
        if style:
            ws.cell(ri, ci).fill = style[0]
            ws.cell(ri, ci).font = style[1]


def _autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 10), 48)


def _setup(ws):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

# ── Output writers ────────────────────────────────────────────────────────────

def _po_out(df_rows: pd.DataFrame) -> pd.DataFrame:
    """Canonical 6-column layout used by all three PO output sheets."""
    out = df_rows[["class", "Item", "Designation", "translation", "db_translation", "Status"]].copy()
    return out.rename(columns={"translation": "PO Translation", "db_translation": "DB Translation"})


def _write_output_po(wb, df: pd.DataFrame):
    """
    PO mode — 3 sheets, all sharing the same 6-column layout:
      class | Item | Designation | PO Translation | DB Translation | Status

      1. Check Results    — NO CHANGE + CHANGED rows (PO translation present, cross-checked vs DB)
      2. Need Translation — NEW ITEM rows (not in DB — paste to Claude)
      3. Ready            — IN DB rows (DB supplies the translation automatically)

    PO Translation is blank on sheets 2/3 when the SPM file carries no French Designation
    for those items — that is expected.  DB Translation is blank on sheet 2 because those
    items are genuinely not yet in the DB.
    """
    # ── Sheet 1: Check Results ────────────────────────────────────
    checked = df[df["translation"].str.strip() != ""].copy()
    checked["_sort"] = checked["Status"].map({CHANGED: 0, NO_CHANGE: 1}).fillna(2)
    checked = checked.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)
    out1 = _po_out(checked)
    ws1 = wb.create_sheet("Check Results")
    ws1.sheet_properties.tabColor = "2E4057"
    _setup(ws1)
    _write_header(ws1, list(out1.columns), _fill("2E4057"))
    _write_rows(ws1, out1)
    _color_status_col(ws1, out1, "Status")
    po_ci = list(out1.columns).index("PO Translation") + 1
    db_ci = list(out1.columns).index("DB Translation") + 1
    for ri in range(2, len(out1) + 2):
        if out1.iloc[ri - 2]["Status"] == CHANGED:
            ws1.cell(ri, po_ci).fill = _fill("FCE4D6")
            ws1.cell(ri, db_ci).fill = _fill("DDEBF7")
    _autofit(ws1)

    # ── Sheet 2: Need Translation ─────────────────────────────────
    out2 = _po_out(df[df["Status"] == NEW_ITEM].copy())
    ws2 = wb.create_sheet("Need Translation")
    ws2.sheet_properties.tabColor = "C55A11"
    _setup(ws2)
    _write_header(ws2, list(out2.columns), _fill("C55A11"))
    _write_rows(ws2, out2)
    _color_status_col(ws2, out2, "Status")
    _autofit(ws2)

    # ── Sheet 3: Ready ────────────────────────────────────────────
    out3 = _po_out(df[df["Status"] == IN_DB].copy())
    ws3 = wb.create_sheet("Ready")
    ws3.sheet_properties.tabColor = "375623"
    _setup(ws3)
    _write_header(ws3, list(out3.columns), _fill("375623"))
    _write_rows(ws3, out3)
    _color_status_col(ws3, out3, "Status")
    _autofit(ws3)


def _write_output_meinv(wb, df: pd.DataFrame):
    """
    MEINV mode — 3 sheets:
      1. Status Check — full audit, all rows
      2. Changed      — translation conflicts
      3. New Items    — not in DB
    """
    # ── Sheet 1: Status Check ─────────────────────────────────────
    out = df[["class", "Item", "translation", "db_translation", "Designation", "Status"]].copy()
    out = out.rename(columns={"translation": "Invoice Translation", "db_translation": "DB Translation"})
    ws = wb.create_sheet("Status Check")
    ws.sheet_properties.tabColor = "2E4057"
    _setup(ws)
    _write_header(ws, list(out.columns), _fill("2E4057"))
    _write_rows(ws, out)
    _color_status_col(ws, out, "Status")
    _autofit(ws)

    # ── Sheet 2: Changed ──────────────────────────────────────────
    rows = df[df["Status"] == CHANGED].copy()
    out2 = pd.DataFrame({
        "class":           rows["class"].values,
        "Item":            rows["Item"].values,
        "Designation":     rows["Designation"].values,
        "New Translation": rows["translation"].values,
        "Old Translation": rows["db_translation"].values,
        "Status":          rows["Status"].values,
    })
    ws2 = wb.create_sheet("Changed")
    ws2.sheet_properties.tabColor = "C55A11"
    _setup(ws2)
    _write_header(ws2, list(out2.columns), _fill("C55A11"))
    _write_rows(ws2, out2)
    _color_status_col(ws2, out2, "Status")
    new_ci = list(out2.columns).index("New Translation") + 1
    old_ci = list(out2.columns).index("Old Translation") + 1
    for ri in range(2, len(out2) + 2):
        ws2.cell(ri, new_ci).fill = _fill("FCE4D6")
        ws2.cell(ri, old_ci).fill = _fill("DDEBF7")
    _autofit(ws2)

    # ── Sheet 3: New Items ────────────────────────────────────────
    rows3 = df[df["Status"] == NEW_ITEM].copy()
    out3 = pd.DataFrame({
        "class":       rows3["class"].values,
        "Item":        rows3["Item"].values,
        "Designation": rows3["Designation"].values,
        "Status":      rows3["Status"].values,
    })
    ws3 = wb.create_sheet("New Items")
    ws3.sheet_properties.tabColor = "375623"
    _setup(ws3)
    _write_header(ws3, list(out3.columns), _fill("375623"))
    _write_rows(ws3, out3)
    _color_status_col(ws3, out3, "Status")
    _autofit(ws3)


def write_output(path: Path, df: pd.DataFrame, mode: str):
    """Dispatch to the correct writer based on mode."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if mode == "PO":
        _write_output_po(wb, df)
    else:
        _write_output_meinv(wb, df)
    wb.save(path)

# ── Main ──────────────────────────────────────────────────────────────────────

def abort(msg: str):
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Translation Checker — Error", msg)
        root.destroy()
    except Exception:
        print(f"\nERROR: {msg}")
    try:
        input("\nPress Enter to close...")
    except (EOFError, KeyboardInterrupt):
        pass
    sys.exit(1)


def main():
    print("=" * 60)
    print("  LG ALGERIA — TRANSLATION CHECKER")
    print("=" * 60)

    # ── Ensure folders exist ──────────────────────────────────────
    INCOMING.mkdir(parents=True, exist_ok=True)
    DB_DIR.mkdir(parents=True, exist_ok=True)

    # ── Find file in incoming/ ────────────────────────────────────
    invoices = sorted(
        [f for f in INCOMING.iterdir() if f.suffix.lower() == ".xlsx" and not f.name.startswith("~$")],
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    if not invoices:
        abort(
            f"No .xlsx file found in:\n  {INCOMING}\n\n"
            f"Drop your PO or MEINV file there and try again."
        )

    invoice_path = invoices[0]
    if len(invoices) > 1:
        print(f"\n  NOTE: {len(invoices)} files in incoming/ — using most recent.")
        print(f"  Others: {[f.name for f in invoices[1:]]}")

    print(f"\n  File     : {invoice_path.name}")
    print()

    # ── Load ──────────────────────────────────────────────────────
    print("  Loading data...")
    try:
        df_invoice, mode = load_invoice(invoice_path)
        df_db            = load_db()
    except Exception as e:
        abort(str(e))

    if len(df_db) == 0:
        print("  WARNING: DB is empty — all items will be classified as NEW ITEM.")

    # ── Classify ──────────────────────────────────────────────────
    print("\n  Classifying rows...")
    df_classified = classify(df_invoice, df_db)

    n_no_change = int((df_classified["Status"] == NO_CHANGE).sum())
    n_in_db     = int((df_classified["Status"] == IN_DB).sum())
    n_changed   = int((df_classified["Status"] == CHANGED).sum())
    n_new       = int((df_classified["Status"] == NEW_ITEM).sum())
    total       = len(df_classified)

    print(f"\n  {'─' * 44}")
    print(f"  Total    : {total} rows")
    if n_no_change: print(f"  {NO_CHANGE}  : {n_no_change}")
    if n_in_db:     print(f"  {IN_DB}       : {n_in_db}  (translation ready)")
    if n_changed:   print(f"  {CHANGED}    : {n_changed}  (needs review)")
    if n_new:       print(f"  {NEW_ITEM}   : {n_new}  (needs translation)")
    print(f"  {'─' * 44}")

    # ── Write output ──────────────────────────────────────────────
    month_dir = OUTPUT_DIR / datetime.now().strftime("%Y-%m")
    month_dir.mkdir(parents=True, exist_ok=True)

    out_name = f"check_{invoice_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    out_path = month_dir / out_name

    print(f"\n  Writing → output/{datetime.now().strftime('%Y-%m')}/{out_name}")

    if mode == "PO":
        print(f"    Sheet 1 — Check Results     ({total} rows)  [{n_changed} changed, {n_no_change} ok, {n_new} new]")
        print(f"    Sheet 2 — Need Translation  ({n_new} rows)  ← paste to Claude")
        print(f"    Sheet 3 — Ready             ({n_in_db} rows)")
    else:
        print(f"    Sheet 1 — Status Check  ({total} rows)")
        print(f"    Sheet 2 — Changed       ({n_changed} rows)")
        print(f"    Sheet 3 — New Items     ({n_new} rows)")

    write_output(out_path, df_classified, mode)

    print(f"\n  [DONE]  {out_path}")

    if mode == "PO":
        print(f"\n  NEXT STEPS:")
        print(f"    1. Open Sheet 2 (Need Translation) — {n_new} items need French customs terms")
        print(f"    2. Paste to Claude with the translation prompt")
        print(f"    3. Save result as xlsx with columns: class | Item | Designation | translation")
        print(f"    4. Drop that file into db/ — done, future runs will find these items")
    else:
        print(f"\n  TIP: Once reviewed, move {invoice_path.name} → db/ to add it to history.")

    try:
        os.startfile(out_path)
    except Exception:
        pass

    try:
        input("\nPress Enter to close...")
    except (EOFError, KeyboardInterrupt):
        pass


if __name__ == "__main__":
    main()
